import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active

ws.title = "First Sheet"

wb.create_sheet(index=1, title="Second Sheet")

ws["A1"] = "Invoice"

# ws['A1'] = Font(name='Times New Roman', size= 24, bold=True, italic=False) #### Value error here
# OR
myfont = Font(name="Times New Roman", size=24, bold=True, italic=False)
ws["A1"].font = myfont

ws["A2"] = "Tires"
ws["A3"] = "Brakes"
ws["A4"] = "Alignment"

ws.merge_cells("A1:B1")

ws["B2"] = 450
ws["B3"] = 225
ws["B4"] = 150

ws["A8"] = "Total"

ws["A8"].font = myfont

ws["B8"] = "=SUM(B2:B4)"

#EXERCISE/PRACTICE

# Read the excel file - 'ProduceReport.xlsx' that you created earlier.
# Write all the contents of this file to 'Second Sheet' in the current workbook
# Display the Grand Total and Average of 'Amt Sold' and 'Total'. At the bottom of the list along with appropriate labels.

write_sheet = wb['Second Sheet']#Writing to 

read_wb = xl.load_workbook('ProduceReport.xlsx') #Reading from
read_ws = read_wb['ProduceReport']

maxC = read_ws.max_column                 #Read Column
maxR = read_ws.max_row                    #Read Row

write_sheet['A1'] = 'Produce'
write_sheet['B1'] = 'Cost Per Pound'
write_sheet['C1'] = 'Amt Sold'
write_sheet['D1'] = 'Total'

write_row = 2
write_colA = 1
write_colB = 2
write_colC = 3
write_colD = 4

for currentrow in read_ws.iter_rows(min_row=2, max_row=maxR,max_col=maxC):
    name = currentrow[0].value
    cost = float(currentrow[1].value)
    amt_sold = float(currentrow[2].value)
    total = float(currentrow[3].value)

    #Writes values to new sheet
    write_sheet.cell(write_row,write_colA).value = name
    write_sheet.cell(write_row,write_colB).value = cost
    write_sheet.cell(write_row,write_colC).value = amt_sold
    write_sheet.cell(write_row,write_colD).value = total

    #Do, otherwise it will be on one line
    write_row += 1
#For Total
summary_row = write_row + 1

write_sheet['B' + str(summary_row)] = 'Total' #Made dynamic
write_sheet['B' + str(summary_row)].font = Font(size=16,bold=True)

write_sheet['C' + str(summary_row)] = '=SUM(C2:C'+str(write_row)+ ')'            # Dynamic, SUM(C2:C42)
write_sheet['D' + str(summary_row)] = '=SUM(D2:D'+str(write_row)+ ')' 

#For Averages
summary_row += 1
write_sheet['B' + str(summary_row)] = 'Average' #Made dynamic
write_sheet['B' + str(summary_row)].font = Font(size=16,bold=True)

write_sheet['C' + str(summary_row)] = '=AVERAGE(C2:C'+str(write_row)+ ')'            # Dynamic, Average(C2:C42)
write_sheet['D' + str(summary_row)] = '=AVERAGE(D2:D'+str(write_row)+ ')'


#Format
write_sheet.column_dimensions['A'].width = 16
write_sheet.column_dimensions['B'].width = 15
write_sheet.column_dimensions['C'].width = 15
write_sheet.column_dimensions['D'].width = 15



for cell in write_sheet["C:C"]:
    cell.number_format = '#,##0'

for cell in write_sheet["D:D"]:
    cell.number_format = u'"$ "#,##0.00'


wb.save("PythontoExcel.xlsx")           #Saving Workbook
