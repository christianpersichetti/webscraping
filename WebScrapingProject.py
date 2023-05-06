# Scrape Crypto Websites for data
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font

# Crypto Webpage
webpage = "https://coinmarketcap.com/"

page = urlopen(webpage)

soup = BeautifulSoup(page, "html.parser")

title = soup.title

# Setting up the Excel Document
wb = xl.Workbook()
ws = wb.active

ws.title = "Crypto Currencies Report"
ws["A1"] = "Name"
ws["B1"] = "Image"
ws["C1"] = "Current Price (USD)"
ws["D1"] = "Percent Change in Past 24 Hours"
ws["E1"] = "Price Change (USD)"

#Previous Prices
previous_price = {"Bitcoin": 0, "Ethereum": 0}

# Webscraping the Website
crypto_rows = soup.findAll("tr")


for x in range (1, 6):
    td = crypto_rows[x].findAll("td")
    name = td[2].text
    image = td[2].find("img")["src"]
    current_price = int(td[3].text.replace("$", "").replace(",", "").replace(".", ""))
    percent_change = int(td[5].text.replace(".", "").replace("%", ""))
    corresponding_price = percent_change * current_price
    #Data into Worksheet
    ws["A" + str(x + 1)] = name
    ws["B" + str(x + 1)] = image
    ws["C" + str(x + 1)] = "$" + format(current_price, ',.2f')
    ws["D" + str(x + 1)] = format(percent_change, '.2f') + '%'
    ws["E" + str(x + 1)] = "$" + format(corresponding_price, ',.2f')


# Dimensions for the Data in the Excel file
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 25
ws.column_dimensions["D"].width = 45
ws.column_dimensions["E"].width = 30

#Font
header_font = Font(size=16, bold=True)
for cell in ws[1:1]:
    cell.font = header_font

#Save Workbook
wb.save("CryptoCurrenciesReport.xlsx")


# Twilio text 
import keys
from twilio.rest import Client

for x in range(1,6):
    td = crypto_rows[x].findAll("td")
    name = td[2].text
    current_price = int(td[3].text.replace("$","").replace(",","").replace(".",""))
    if name == "Bitcoin" and (current_price - previous_price['Bitcoin']) <= 5:
        message = f"Bitcoin price has changed by ${(current_price - previous_price['Bitcoin'])}"
        #Updating the price
        previous_price[name] = current_price
        
    elif name == "Ethereum" and (current_price - previous_price['Ethereum']) <= 5:
        message = f"Ethereum price has changed by ${(current_price - previous_price['Ethereum'])}"
        
        #Updating the price
        previous_price[name] = current_price

    else:
        message = f"There was no significant price change for Bitcoin or Ethereum"

    client = Client(keys.accountSID, keys.auth_token)
    TwilioNumber = "+7207800915"
    myNumber = "+7202916302"
    textmessage = client.messages.create(to=myNumber, from_=TwilioNumber, body=message)


